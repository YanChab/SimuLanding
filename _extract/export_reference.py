import openpyxl, csv, os, json
xlsm = r"c:\Users\ychab\Programmation\SimuLanding\DROSIM_SA61-_#Simulation drop test avion complet.xlsm"
outdir = r"c:\Users\ychab\Programmation\SimuLanding\_extract\reference"
os.makedirs(outdir, exist_ok=True)
wb = openpyxl.load_workbook(xlsm, data_only=True, read_only=True)

# Export Results MLG to CSV (header row 2, data from row 3)
ws = wb["Results MLG"]
ncol = ws.max_column
with open(os.path.join(outdir, "Results_MLG.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    header = [ws.cell(row=2, column=c).value for c in range(2, ncol+1)]
    w.writerow(header)
    for r in range(3, ws.max_row+1):
        row = [ws.cell(row=r, column=c).value for c in range(2, ncol+1)]
        if all(v is None for v in row):
            continue
        w.writerow(row)
print("Wrote Results_MLG.csv")
wb.close()
print("DONE")
