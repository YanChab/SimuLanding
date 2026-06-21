import csv
import os
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSM = ROOT / "DROSIM_SA61-_#Simulation drop test avion complet.xlsm"
OUTDIR = ROOT / "_extract" / "reference"
OUTDIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(XLSM, data_only=True, read_only=True)
ws = wb["Results NLG"]

ncol = ws.max_column
out_csv = OUTDIR / "Results_NLG.csv"
with out_csv.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    header = [ws.cell(row=2, column=c).value for c in range(2, ncol + 1)]
    w.writerow(header)
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(2, ncol + 1)]
        if all(v is None for v in row):
            continue
        w.writerow(row)

wb.close()
print(f"Wrote {out_csv}")
