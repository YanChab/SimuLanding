import openpyxl
xlsm = r"c:\Users\ychab\Programmation\SimuLanding\DROSIM_SA61-_#Simulation drop test avion complet.xlsm"
wb = openpyxl.load_workbook(xlsm, data_only=True, read_only=True)

# Results MLG : headers (row 1..3) + first 6 data rows, cols A..AP
ws = wb["Results MLG"]
print("===== Results MLG dims:", ws.max_row, "x", ws.max_column)
for r in range(1, 9):
    parts = []
    for c in range(1, min(ws.max_column, 45)+1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            parts.append(f"{openpyxl.utils.get_column_letter(c)}{r}={v!r}")
    if parts:
        print(" | ".join(parts))

# tyre table + mu table full (MLG sheet B40..C90)
ws2 = wb["MLG"]
print("\n===== MLG tyre table B40..C47 and mu table B64..C90 =====")
print("TYRE (defl mm, load kN):")
for r in range(40, 48):
    b = ws2.cell(row=r, column=2).value
    c = ws2.cell(row=r, column=3).value
    if b is not None:
        print(f"  {b!r}, {c!r}")
print("MU/SLIP (slip, mu):")
for r in range(64, 95):
    b = ws2.cell(row=r, column=2).value
    c = ws2.cell(row=r, column=3).value
    if b is not None or c is not None:
        print(f"  {b!r}, {c!r}")
wb.close()
print("DONE")
