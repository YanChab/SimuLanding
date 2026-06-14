import openpyxl, json, os
xlsm = r"c:\Users\ychab\Programmation\SimuLanding\DROSIM_SA61-_#Simulation drop test avion complet.xlsm"
wb = openpyxl.load_workbook(xlsm, data_only=True, read_only=True)
print("SHEETS:", wb.sheetnames)

def dump_grid(ws, rmax, cmax):
    rows = []
    for r in range(1, rmax+1):
        cells = []
        for c in range(1, cmax+1):
            v = ws.cell(row=r, column=c).value
            cells.append(v)
        rows.append((r, cells))
    return rows

# ---- MLG sheet : input zone (rows 1..70, cols 1..20) ----
ws = wb["MLG"]
print("\n===== MLG (input) rows 1..70, cols A..T =====")
for r, cells in dump_grid(ws, 70, 20):
    # only print non-empty rows
    if any(v is not None and str(v).strip() != "" for v in cells):
        # show col letter:value for non-empty
        parts = []
        for ci, v in enumerate(cells, start=1):
            if v is not None and str(v).strip() != "":
                parts.append(f"{openpyxl.utils.get_column_letter(ci)}{r}={v!r}")
        print("  " + " | ".join(parts))

wb.close()
print("\nDONE")
