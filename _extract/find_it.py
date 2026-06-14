import openpyxl
wb = openpyxl.load_workbook(r'DROSIM_SA61-_#Simulation drop test avion complet.xlsm', data_only=True)
ws = wb['MLG']
lines = []
# Recherche des libellés It / pas de temps / temps simu autour des cellules de paramètres
for r in range(1, 60):
    for cc in range(1, 14):
        v = ws.cell(r, cc).value
        if isinstance(v, str) and any(s in v.lower() for s in ['it', 'pas', 'temps', 'incr', 'step', 'simu']):
            lines.append(f"R{r}C{cc} '{v}' -> R{r}C{cc+1}={ws.cell(r, cc+1).value} R{r}C{cc+2}={ws.cell(r, cc+2).value}")
open(r'_extract\it_out.txt', 'w', encoding='utf-8').write('\n'.join(lines))
print('done', len(lines))
